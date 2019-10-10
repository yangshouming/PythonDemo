'''
@Description: In User Settings Edit
@Author: your name
@Date: 2019-10-08 11:28:42
@LastEditTime: 2019-10-09 15:42:04
@LastEditors: Please set LastEditors
'''

import os
import sys
from openpyxl import load_workbook
# 读取路径
book = load_workbook(filename=r"excel2tabel.xlsx")
# 读取名字为Sheet1的表
# sheet = book.get_sheet_by_name("Sheet1")
sheet = book["Sheet1"]
# 用于存储数据的数组
data = []

row_num = 1
while row_num <= book.active.max_row:
    col_num = 1
    while col_num <= book.active.max_column:
        data.append(sheet.cell(row=row_num, column=col_num).value)
        col_num = col_num + 1
    row_num = row_num + 1

print('max_row', book.active.max_row)
print('max_column', book.active.max_column)
print('file save')

myfile1 = open("excel2table.txt", 'w')
myfile1.write(str(data))
